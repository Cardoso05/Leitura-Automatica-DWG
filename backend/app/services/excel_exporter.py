from io import BytesIO
from typing import Dict, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def build_excel(
    items: Iterable[dict], summary: Dict[str, float], metadata: Dict
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quantitativo"

    headers = [
        "Disciplina",
        "Categoria",
        "Descrição",
        "Unidade",
        "Quantidade",
        "Layer",
        "Bloco",
    ]
    ws.append(headers)
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold
        ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center")

    for item in items:
        ws.append(
            [
                item.get("discipline"),
                item.get("category"),
                item.get("description"),
                item.get("unit"),
                round(item.get("quantity", 0.0), 2),
                item.get("layer"),
                item.get("block_name"),
            ]
        )

    summary_ws = wb.create_sheet("Resumo")
    summary_ws.append(["Disciplina", "Total"])
    summary_ws["A1"].font = bold
    summary_ws["B1"].font = bold
    for discipline, total in summary.items():
        summary_ws.append([discipline, round(total, 2)])

    meta_ws = wb.create_sheet("Metadata")
    for key, value in metadata.items():
        meta_ws.append([key, value])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
